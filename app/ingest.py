"""Excel ingestion pipeline for Farlight K193 exports (17-column format).

Source: cod-game-tools.farlightgames.com/topn (official beta portal).

The portal lets users tick which columns to include in the export.
This parser is therefore tolerant: any subset of known columns is accepted,
provided the minimal identity columns are present (character_id, name, power,
merits_total).
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from .db import get_session
from .models import Member, Season, Setting, Snapshot, Stat
from .security import require_ingest_token

router = APIRouter(prefix="/api", tags=["ingest"])


HEADER_MAP = {
    "rang": "rank",
    "identifiant du personnage": "character_id",
    "nom du personnage": "current_name",
    "puissance actuelle": "power",
    "plus haute puissance historique": "peak_power",
    "morts (t4/t5)": "deaths_t45",
    "merites totaux": "merits_total",
    "infanterie uniquement": "merits_infantry",
    "cavalerie uniquement": "merits_cavalry",
    "tireurs d'elite uniquement": "merits_archers",
    "unites magiques uniquement": "merits_magic",
    "autres merites": "merits_other",
    "guerison (t4/t5)": "healing_t45",
    "dons de l'alliance": "alliance_donations",
    "temps de construction": "build_time",
    "temps de destruction": "destruction_time",
    "victoires lors de raids de behemoth": "behemoth_victories",
    "recolte": "harvest",
}

REQUIRED_FIELDS = {"character_id", "current_name", "power", "merits_total"}

_FILENAME_DATES_RE = re.compile(r"(\d{4}-\d{2}-\d{2})[_-](\d{4}-\d{2}-\d{2})")


def _normalize(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(" ", "").replace(",", "")))
    except (ValueError, TypeError):
        return 0


def _extract_dates_from_filename(filename: str) -> tuple[date, date]:
    m = _FILENAME_DATES_RE.search(filename)
    if not m:
        today = date.today()
        return today, today
    return date.fromisoformat(m.group(1)), date.fromisoformat(m.group(2))


def _get_setting_bool(session: Session, key: str, default: bool) -> bool:
    setting = session.execute(select(Setting).where(Setting.key == key)).scalar_one_or_none()
    if setting is None:
        return default
    return bool(setting.value.get("v", default))


def _get_active_season(session: Session) -> Season:
    season = session.execute(
        select(Season).where(Season.is_active.is_(True))
    ).scalar_one_or_none()
    if season is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No active season. Seed one via init_db.",
        )
    return season


def _detect_overlap(session: Session, season_id: int, date_start: date, date_end: date):
    candidates = session.execute(
        select(Snapshot).where(Snapshot.season_id == season_id)
    ).scalars().all()
    for c in candidates:
        if max(c.date_start, date_start) <= min(c.date_end, date_end):
            return c
    return None


@router.post("/ingest", status_code=status.HTTP_201_CREATED)
async def ingest_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    _: None = Depends(require_ingest_token),
):
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot open as Excel: {exc}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Empty file.")

    columns = []
    seen_fields = set()
    for cell in header_row:
        key = _normalize(cell)
        field = HEADER_MAP.get(key)
        columns.append(field)
        if field:
            seen_fields.add(field)

    missing = REQUIRED_FIELDS - seen_fields
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {sorted(missing)}",
        )

    filename = file.filename or "unknown.xlsx"
    date_start, date_end = _extract_dates_from_filename(filename)

    existing = session.execute(
        select(Snapshot).where(
            and_(
                Snapshot.source_filename == filename,
                Snapshot.date_start == date_start,
                Snapshot.date_end == date_end,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Snapshot already ingested (id={existing.id}).",
        )

    season = _get_active_season(session)
    if _get_setting_bool(session, "ingest.reject_overlapping_periods", default=True):
        overlap = _detect_overlap(session, season.id, date_start, date_end)
        if overlap:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Date range overlaps existing snapshot id={overlap.id} "
                    f"({overlap.date_start} to {overlap.date_end})."
                ),
            )

    parsed_rows = []
    skipped = 0
    for row in rows_iter:
        if row is None or all(c is None or c == "" for c in row):
            continue
        record = {}
        for col_field, value in zip(columns, row):
            if col_field is None:
                continue
            record[col_field] = value
        if not record.get("character_id") or not record.get("current_name"):
            skipped += 1
            continue
        try:
            record["character_id"] = int(record["character_id"])
        except (ValueError, TypeError):
            skipped += 1
            continue
        parsed_rows.append(record)

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No valid data rows.")

    snapshot = Snapshot(
        season_id=season.id,
        date_start=date_start,
        date_end=date_end,
        source_filename=filename,
        row_count=len(parsed_rows),
        ingested_at=datetime.utcnow(),
        ingested_by="api",
    )
    session.add(snapshot)
    session.flush()

    members_created = 0
    members_updated = 0
    now = datetime.utcnow()

    char_ids = [r["character_id"] for r in parsed_rows]
    existing_members = {
        m.character_id: m
        for m in session.execute(
            select(Member).where(Member.character_id.in_(char_ids))
        ).scalars()
    }

    stat_objects = []
    for r in parsed_rows:
        cid = r["character_id"]
        name = str(r["current_name"]).strip()

        member = existing_members.get(cid)
        if member is None:
            member = Member(
                character_id=cid,
                current_name=name,
                in_alliance=False,
                troop_tier="unknown",
                first_seen_at=now,
                last_seen_at=now,
            )
            session.add(member)
            members_created += 1
        else:
            member.current_name = name
            member.last_seen_at = now
            members_updated += 1

        stat_objects.append(
            Stat(
                snapshot_id=snapshot.id,
                character_id=cid,
                rank=_parse_int(r.get("rank")) or None,
                power=_parse_int(r.get("power")),
                peak_power=_parse_int(r.get("peak_power")) or None,
                deaths_t45=_parse_int(r.get("deaths_t45")),
                merits_total=_parse_int(r.get("merits_total")),
                merits_infantry=_parse_int(r.get("merits_infantry")),
                merits_cavalry=_parse_int(r.get("merits_cavalry")),
                merits_archers=_parse_int(r.get("merits_archers")),
                merits_magic=_parse_int(r.get("merits_magic")),
                merits_other=_parse_int(r.get("merits_other")),
                healing_t45=_parse_int(r.get("healing_t45")),
                alliance_donations=_parse_int(r.get("alliance_donations")),
                build_time=_parse_int(r.get("build_time")),
                destruction_time=_parse_int(r.get("destruction_time")),
                behemoth_victories=_parse_int(r.get("behemoth_victories")),
                harvest=_parse_int(r.get("harvest")),
            )
        )

    session.add_all(stat_objects)
    session.commit()

    return {
        "snapshot_id": snapshot.id,
        "season_id": season.id,
        "date_start": str(date_start),
        "date_end": str(date_end),
        "rows_parsed": len(parsed_rows),
        "rows_skipped": skipped,
        "members_created": members_created,
        "members_updated": members_updated,
        "columns_detected": sorted(seen_fields),
    }
